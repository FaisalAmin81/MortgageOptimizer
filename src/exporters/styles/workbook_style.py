from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
)

from openpyxl.styles.colors import Color


class WorkbookStyle:

    # -------------------------------------------------
    # Fonts
    # -------------------------------------------------

    TITLE_FONT = Font(
        name="Calibri",
        size=18,
        bold=True,
        color="1F1F1F",
    )

    HEADER_FONT = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF",
    )

    NORMAL_FONT = Font(
        name="Calibri",
        size=11,
    )

    BOLD_FONT = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    # -------------------------------------------------
    # Colors
    # -------------------------------------------------

    HEADER_FILL = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    LIGHT_BLUE_FILL = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    LIGHT_GRAY_FILL = PatternFill(
        "solid",
        fgColor="F4F4F4",
    )

    GREEN_FILL = PatternFill(
        "solid",
        fgColor="C6EFCE",
    )

    RED_FILL = PatternFill(
        "solid",
        fgColor="FFC7CE",
    )

    # -------------------------------------------------
    # Borders
    # -------------------------------------------------

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -------------------------------------------------
    # Alignment
    # -------------------------------------------------

    LEFT = Alignment(horizontal="left")

    CENTER = Alignment(horizontal="center")

    RIGHT = Alignment(horizontal="right")

    # -------------------------------------------------
    # Number Formats
    # -------------------------------------------------

    CURRENCY = "#,##0.00"

    PERCENT = "0.00%"

    INTEGER = "#,##0"
