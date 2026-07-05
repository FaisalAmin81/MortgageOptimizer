from openpyxl import Workbook

from src.exporters.sheets.summary_sheet import SummarySheet
from src.models.simulation_result import SimulationResult


class ExcelExporter:
    """
    Exports simulation results to Excel.
    """

    @staticmethod
    def export(
        result: SimulationResult,
        filename: str = "MortgageReport.xlsx",
    ) -> None:

        workbook = Workbook()

        workbook.remove(workbook.active)

        workbook.create_sheet("Summary")
        workbook.create_sheet("Loan Schedule")
        workbook.create_sheet("Monthly Snapshots")
        workbook.create_sheet("Savings")
        workbook.create_sheet("Settlements")

        SummarySheet.build(
            workbook,
            result,
        )

        workbook.save(filename)
